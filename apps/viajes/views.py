from django.shortcuts import render, redirect, get_list_or_404
from django.http import HttpResponse
from django.views.generic import CreateView, View, ListView, TemplateView
from django.core import serializers
from django.contrib import messages
import pytz

# from django.contrib.auth.models import User
# ya q creamos un modelo de usuario personalidado, debemos llamar a ese modelo envez de llamar al modelo del Framework
from apps.account.models import Account as User
from datetime import datetime

from django.utils.decorators import method_decorator
from apps.home.decoradores_viaje import premiso_transportista


from apps.viajes.models import Viaje, ViajeAdministrativo
from apps.viajes.forms import viaje_form, viaje_admin_form
from apps.pasajero.models import Pasajero
from apps.ruta.models import Ruta
from apps.vendor.models import Vendor

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font


from datetime import datetime
import random

from django.utils.decorators import method_decorator
from apps.home.decoradores_viaje import premiso_admin, premiso_transportista


# numero_viaje: esta variable de se uililizara para crear el codigo de Viaje, este sera basado en la informacion del usuario y la fecha para asi generar un codigo unico
# el numero de viaje estara compuesto por "TPT-"+ 3 primeras letras de nombre de usuario usuario en mayuscula + dia + mes + año + hora actual al momento de hacer loguin + "-" + un numero random entre 100 y 10000 "
# se toman la hora actual y la fecha, esta se usara para crear la variable de sesion para el numero de viaje

fecha_tiempo = datetime.now()
hora_actual = str(fecha_tiempo.hour)
minuto_actual = str(fecha_tiempo.minute)
dia = str(fecha_tiempo.day)
mes = str(fecha_tiempo.month)
ano = str(fecha_tiempo.year)
numero_random = random.randint(100, 10000)

# aqui se crea el numero de viaje y se almacena en una variable de sesion, para posteriormente se urilizado
# este numero se usara para cosas como, indicar el numero de viaje en la base de datos, filtrar y organizar los viajes por numero de viajes
# mostrar los viajes de la sesion activa en el template listar_viaje_en_curso.html
@premiso_transportista
def generar_viaje(request):
     #si no es un Supe usuario o un Transportista la app manda un mensaje de eeor y lo envia al home
    if not request.user.is_transportista and not request.user.is_superuser:
        messages.success(request, "No tiene permisos Para Acceder a seccion")
        return redirect("home")

    tipo_viaje = request.POST["tipo_viaje"]
    hora_viaje = request.POST["hora_viaje"]
    try:
        nombre_usuario = request.user.username
        request.session["numero_viaje"] = (
            "TPG-"
            + str(nombre_usuario[:3]).upper()
            + dia
            + mes
            + ano
            + "-"
            + str(random.randint(1, 20000))
        )
        request.session["hora"] = hora_actual
        request.session["hora_viaje"] = hora_viaje
        request.session["tipo_viaje"] = tipo_viaje
        return redirect("crear_viaje")
    except KeyError:
        pass

@method_decorator(premiso_transportista,name='dispatch')
class agregar_viaje(CreateView):
    form_class = viaje_form
    template_name = "viajes_templates/viaje_form.html"

    
    def get(self, request, id_escaneado):
        

        fecha_tiempo = datetime.now()

        # formatear hora a 12 horas
        hora_actual = fecha_tiempo.hour
        meridiano = fecha_tiempo.strftime("%p")

        usuario_logueado = User.objects.get(username=request.user.username)

        # converir a formato de 12 horas logica

        # confirmar AM PM
        if hora_actual < 11 or hora_actual == 0:
            meridiano = "AM"
        else:
            meridiano = "PM"

        # pasajero_escaneado = serializers.serialize('json', Pasajero.objects.filter(id_pasajero=id_escaneado))

        # con el escaner se bueca el pasajero este Id es pasado por la URL, y capturado como parametro en la funcion agregar_viaje
        # luego este se usa para filtrar la lista de Pasajeros , ya q es una llave primaria deberia devolver solo un pasajero y este seria el 0, primero y unico
        # una vez se obtiene ese pasajero se puede acceder a cada atributo del pasajero y asignarle el valor a cada uno de los campos del modelo Viaje para agregar pasajeros a los viajes.
        try:
            pasajero_escaneado = Pasajero.objects.filter(id_pasajero=id_escaneado)
        

            id_pasajero = pasajero_escaneado[0].id_pasajero
            nombre = pasajero_escaneado[0].nombre
            apellido = pasajero_escaneado[0].apellido
            direccion = pasajero_escaneado[0].direccion
            telefono = pasajero_escaneado[0].telefono
            ruta_pasajero = pasajero_escaneado[0].Ruta
            cuenta = pasajero_escaneado[0].cuenta
            site = pasajero_escaneado[0].site
            hora_entrada = pasajero_escaneado[0].hora_entrada
            hora_salida = pasajero_escaneado[0].hora_salida

        except (IndexError):
            messages.success(request, "No se encontro el pasajero o el mismo no existe")
            return render(request,'mensaje.html')
        except Exception as e:
            messages.success(request, "error: "+ str(e))
            return render(request,'mensaje.html')

        # para confirmar que el Pasajero esta dentro de su rango de hora para abordar el transporte
        # confirmamos que la hora actual sea igual a la hora de entrada, o que la hora actual sea igual a la hora de salida
        # de este modo la aplicacion crearia el viaje y dejaria abordar al pasajero en cualquiera de las dos horas

        if hora_actual == hora_entrada.hour or hora_actual == hora_salida.hour or hora_actual < hora_salida.hour:

            # se valida si el pasajero existe en el viaje actual, se toma el numero de viaje q es una variable de sesion y se toma el id escaneado
            # si se detecta el id_pasajero y el numero_viaje en el mismo objero del qeryset no se agrega el pasajero a viaje puesto que ese pasajero ya esta en ese viaje
            qs = Viaje.objects.filter(
                numero_viaje=request.session["numero_viaje"], id_pasajero=id_escaneado
            )
            if qs:
                messages.success(request, "El pasajero ya existe en este viaje")
                return render(request, "mensaje.html")
                # print("ya Existe el pasajero")
            else:

                V = Viaje(
                    numero_viaje=request.session["numero_viaje"],
                    id_viaje=None,
                    transportista=usuario_logueado,
                    fecha_viaje=fecha_tiempo,
                    hora_viaje=request.session["hora_viaje"],
                    tipo_viaje=request.session["tipo_viaje"],
                    hora_entrada=hora_entrada,
                    hora_salida=hora_salida,
                    id_pasajero=id_pasajero,
                    nombre_pasajero=nombre,
                    apellido_pasajero=apellido,
                    campaña_pasajero=cuenta,
                    site_pasajero=site,
                    ruta_pasajero=ruta_pasajero,
                    direccion_pasajero=direccion,
                    # cantidad_pasajeros =5,
                    # precio_viaje =5,
                )
                if V:
                    V.save()
                    messages.success(request, "Se agrego el pasajero correctamente")
                    return render(request, "mensaje.html", context={"form": V})
        else:
            messages.success(
                request,
                "El Pasajero esta Fuera de su hora, si desea agregaro debe elegir la opcion manual",
            )
            return render(request, "mensaje.html")

@method_decorator(premiso_transportista,name='dispatch')
class crear_viaje(CreateView):
    model = Viaje
    form_class = viaje_form
    template_name = "viajes_templates/viaje_form.html"
    success_url = "listar_pasajeros"

    def post(self, request):
        

        # se valida si el pasajero existe en el viaje actual, se toma el numero de viaje q es una variable de sesion y se toma el id escaneado
        # si se detecta el id_pasajero y el numero_viaje en el mismo objero del qeryset no se agrega el pasajero a viaje puesto que ese pasajero ya esta en ese viaje
        qs = Viaje.objects.filter(
            numero_viaje=request.session["numero_viaje"],
            id_pasajero=request.POST["id_pasajero"],
        )
        if qs:
            print("pasajero existe en este viaje")
            messages.warning(request, "El pasajero ya existe en este viaje")
            return redirect("crear_viaje")
        else:

            V = Viaje(
                numero_viaje=request.session["numero_viaje"],
                id_viaje=None,
                transportista=request.POST["transportista"],
                fecha_viaje=request.POST["fecha_viaje"],
                hora_viaje=request.session["hora_viaje"],
                tipo_viaje=request.session["tipo_viaje"],
                hora_entrada=request.POST["hora_entrada"],
                hora_salida=request.POST["hora_salida"],
                id_pasajero=request.POST["id_pasajero"],
                nombre_pasajero=request.POST["nombre_pasajero"],
                apellido_pasajero=request.POST["apellido_pasajero"],
                campaña_pasajero=request.POST["campaña_pasajero"],
                site_pasajero=request.POST["site_pasajero"],
                ruta_pasajero=request.POST["ruta_pasajero"],
                direccion_pasajero=request.POST["direccion_pasajero"],
                excepcion=request.POST["excepcion"],
                razon_excepcion=request.POST["razon_excepcion"],
                # cantidad_pasajeros =5,
                # precio_viaje =5,
            )
            if V:
                V.save()
                messages.success(request, "Se agrego el pasajero correctamente")
                return redirect("crear_viaje")
            else:
                messages.success(request, "Ha Habido un error Creando el Viaje")

                return redirect("crear_viaje")
@method_decorator(premiso_transportista,name='dispatch')
class agregar_viaje_manual(View):
    model = Viaje
    form_class = viaje_form
    # template_name = 'agregar_viaje_manual.html'
    success_url = "#"

    # obtener datos de usuario para llenar formulario automaticamente
    def get(self, request, id_ingresado):
         
        try:
            # pasajero_escaneado = serializers.serialize('json', Pasajero.objects.filter(id_pasajero=id_escaneado))

            # con el escaner se bueca el pasajero este Id es pasado por la URL, y capturado como parametro en la funcion agregar_viaje
            # luego este se usa para filtrar la lista de Pasajeros , ya q es una llave primaria deberia devolver solo un pasajero y este seria el 0, primero y unico
            # una vez se obtiene ese pasajero se puede acceder a cada atributo del pasajero y asignarle el valor a cada uno de los campos del modelo Viaje para agregar pasajeros a los viajes.

            pasajero_escaneado = Pasajero.objects.filter(id_pasajero=id_ingresado)
            # strftime año mes dia %Y-%m-%d  devuelve hora mintu segundo %H:%M:%S
            fecha_tiempo = datetime.now().strftime("%Y-%m-%d")
            usuario_logueado = User.objects.get(username=request.user.username)

            id_pasajero = pasajero_escaneado[0].id_pasajero
            nombre = pasajero_escaneado[0].nombre
            apellido = pasajero_escaneado[0].apellido
            direccion = pasajero_escaneado[0].direccion
            telefono = pasajero_escaneado[0].telefono
            ruta_pasajero = pasajero_escaneado[0].Ruta
            cuenta = pasajero_escaneado[0].cuenta
            site = pasajero_escaneado[0].site
            hora_entrada = pasajero_escaneado[0].hora_entrada
            hora_salida = pasajero_escaneado[0].hora_salida
        except (IndexError):
            messages.success(request, "No se encontro el pasajero o el mismo no existe")
            return render(request,'mensaje.html')
        except Exception as e:
            messages.success(request, "error: "+ str(e))
            return render(request,'mensaje.html')
            
        context = {
            "numero_viaje": request.session["numero_viaje"],
            "id_viaje": None,
            "transportista": usuario_logueado,
            "fecha_viaje": fecha_tiempo,
            "hora_entrada": hora_entrada,
            "hora_salida": hora_salida,
            "id_pasajero": id_pasajero,
            "nombre_pasajero": nombre,
            "apellido_pasajero": apellido,
            "campaña_pasajero": cuenta,
            "site_pasajero": site,
            "ruta_pasajero": ruta_pasajero,
            "direccion_pasajero": direccion,
        }

        return render(request, "viajes_templates/agregar_viaje_manual.html", context)


# esta funcion devuelve los viajes en curso, tomando como filtro el numero de viaje que que esta en el momento
@premiso_transportista
def listar_viaje_en_curso(request):

    try:
        
        qs = Viaje.objects.filter(numero_viaje=request.session["numero_viaje"]).order_by('-fecha_viaje')

        # contar el numero de pasajeros en este queryset
        cantidad_pasajeros = qs.count()

        # definir el precio del viaje
        # viajes de 1 a 3 personas paga 12 dolares
        # viajes de 4 a 15 personas paga 26 dolares

        if cantidad_pasajeros >= 1 and cantidad_pasajeros < 4:
            precio_viaje = 12

        if cantidad_pasajeros >= 4 and cantidad_pasajeros < 16:
            precio_viaje = 26

        if cantidad_pasajeros < 1:
            precio_viaje = 0

        # calcular el precio por cada pasajero
        if cantidad_pasajeros > 0:
            precio_por_pasajero = precio_viaje / cantidad_pasajeros
        else:
            precio_por_pasajero = 0

        context = {
            "object_list": qs,
            "cantidad_pasajeros": cantidad_pasajeros,
            "precio_viaje": precio_viaje,
            "precio_por_pasajero": precio_por_pasajero,
        }

        return render(request, "viajes_templates/listar_viaje_en_curso.html", context)
    except Exception as e:
        messages.success(request, str(e) +' '+ "Debe generar un dumero de viaje primero")
        return redirect('home')


# esta funcion devuelve los viajes que se han realizado por el transportista logueado al momento
@premiso_transportista
def mis_viajes(request):
 

    numero_exacto_viaje = request.GET.get("dato")
    qs = Viaje.objects.filter(transportista=request.user.username).order_by('-fecha_viaje')

    if numero_exacto_viaje != "" and numero_exacto_viaje is not None:
        qs = qs.filter(
            transportista=request.user.username,
            numero_viaje__icontains=numero_exacto_viaje,
        ).order_by('-fecha_viaje')

    context = {
        "object_list": qs,
    }

    return render(request, "viajes_templates/listar_viajes_transportista_logueado.html", context)


@method_decorator(premiso_transportista,name='dispatch')
class reporte_viaje_excel(TemplateView):

    # Usamos el método get para generar el archivo excel
    @premiso_transportista
    def get(self, request, *args, **kwargs):
         #si no es un Supe usuario o un Transportista la app manda un mensaje de eeor y lo envia al home
        if not request.user.is_transportista and not request.user.is_superuser:
            messages.success(request, "No tiene permisos Para Acceder a seccion")
            return redirect("home")

        # Obtenemos todas las personas de nuestra base de datos
        viajes = Viaje.objects.all()
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        # Aplicamos estilo a la columna B1
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B2"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["B2"].fill = PatternFill(
            start_color="66FFCC", end_color="66FFCC", fill_type="solid"
        )
        ws["B2"].font = Font(name="Calibi", size=16, bold=True)
        ws["B2"] = "REPORTE DE VIAJES"
        # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells("B2:O2")
        ws.row_dimensions[2].height = 25

        # definimos el tamaño de cada colunma algunas las dejamos en automatico
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["d"].auto_size = True
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].auto_size = True
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].auto_size = True
        ws.column_dimensions["J"].auto_size = True
        ws.column_dimensions["K"].auto_size = True
        ws.column_dimensions["L"].auto_size = True
        ws.column_dimensions["M"].width = 15
        ws.column_dimensions["N"].width = 30
        ws.column_dimensions["O"].width = 30

        # Creamos los encabezados desde la celda B3 hasta la E3
        ws["B5"] = "Id"
        ws["C5"] = "Numero de viaje"
        ws["D5"] = "Transportista"
        ws["E5"] = "Fecha"
        ws["F5"] = "Hora"
        ws["G5"] = "Tipo"
        ws["H5"] = "Id"
        ws["I5"] = "Nombre"
        ws["J5"] = "Apellido"
        ws["K5"] = "Campaña"
        ws["L5"] = "Site"
        ws["M5"] = "Ruta"
        ws["N5"] = "Direccion"
        ws["O5"] = "Excepcion"
        # ws['P3']='Hora de entrada del pasajero'
        # ws['P3']='Hora de salida del pasajero'

        # damos Formato a cada una de las celdas utilizadas
        ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["B5"].font = Font(name="Arial", size=10, bold=True)
        ws["B5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["C5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["C5"].font = Font(name="Arial", size=10, bold=True)
        ws["C5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["D5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["D5"].font = Font(name="Arial", size=10, bold=True)
        ws["D5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["E5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["E5"].font = Font(name="Arial", size=10, bold=True)
        ws["E5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["F5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["F5"].font = Font(name="Arial", size=10, bold=True)
        ws["F5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["G5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["G5"].font = Font(name="Arial", size=10, bold=True)
        ws["G5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["H5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["H5"].font = Font(name="Arial", size=10, bold=True)
        ws["H5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["I5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["I5"].font = Font(name="Arial", size=10, bold=True)
        ws["I5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["J5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["J5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["J5"].font = Font(name="Arial", size=10, bold=True)
        ws["J5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["K5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["K5"].font = Font(name="Arial", size=10, bold=True)
        ws["K5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["L5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["L5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["L5"].font = Font(name="Arial", size=10, bold=True)
        ws["L5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["M5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["M5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["M5"].font = Font(name="Arial", size=10, bold=True)
        ws["M5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["N5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["N5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["N5"].font = Font(name="Arial", size=10, bold=True)
        ws["N5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["O5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["O5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["O5"].font = Font(name="Arial", size=10, bold=True)
        ws["O5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        cont = 6
        # Recorremos el conjunto de viajes y vamos escribiendo cada uno de los datos en las celdas y tambien se le asignan su respectivos estilos
        for viaje in viajes:
            ws.cell(row=cont, column=2).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=2).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=2).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=2).value = viaje.id_viaje

            ws.cell(row=cont, column=3).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=3).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=3).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=3).value = viaje.numero_viaje

            ws.cell(row=cont, column=4).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=4).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=4).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=4).value = viaje.transportista

            ws.cell(row=cont, column=5).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=5).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=5).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=5).value = viaje.fecha_viaje

            ws.cell(row=cont, column=6).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=6).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=6).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=6).value = viaje.hora_viaje

            ws.cell(row=cont, column=7).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=7).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=7).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=7).value = viaje.tipo_viaje

            ws.cell(row=cont, column=8).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=8).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=8).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=8).value = viaje.id_pasajero

            ws.cell(row=cont, column=9).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=9).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=9).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=9).value = viaje.nombre_pasajero

            ws.cell(row=cont, column=10).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=10).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=10).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=10).value = viaje.apellido_pasajero

            ws.cell(row=cont, column=11).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=11).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=11).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=11).value = viaje.campaña_pasajero

            ws.cell(row=cont, column=12).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=12).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=12).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=12).value = viaje.site_pasajero

            ws.cell(row=cont, column=13).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=13).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=13).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=13).value = viaje.ruta_pasajero

            ws.cell(row=cont, column=14).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=14).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=14).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=14).value = viaje.direccion_pasajero

            ws.cell(row=cont, column=15).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=15).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=15).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=15).value = viaje.razon_excepcion

            # ws.cell(row = cont, column =17).value = viaje.hora_entrada
            # ws.cell(row = cont, column =18).value = viaje.hora_salida
            cont = cont + 1
        # Establecemos el nombre del archivo este estara compuesto por la palabra TPGP-REPORT- +dia + mes+ ano +horaactual+minutoactual
        nombre_archivo = (
            "TPGO-REPORT-" + dia + mes + ano + hora_actual + minuto_actual + ".xlsx"
        )
        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response





@premiso_admin
def generar_viaje_admin(request):
     #si no es un Supe usuario o un Transportista la app manda un mensaje de eeor y lo envia al home
   

    try:
        nombre_usuario = request.user.username
        request.session["numero_viaje_admin"] = (
            "TPGAD-"
            + str(nombre_usuario[:3]).upper()
            + dia
            + mes
            + ano
            + "-"
            + str(random.randint(1, 20000))
        )
        request.session["hora"] = hora_actual
        return redirect("crear_viaje_admin")
    except KeyError:
        pass
    

#con method decorator se pueden poner decoradores sobre las clases, y hay q soobrescribir el nombre dispacth
@method_decorator(premiso_admin,name='dispatch')
class crear_viaje_admin(CreateView):
    model = ViajeAdministrativo
    template_name = "viajes_templates/viaje_admin_form.html"
    form_class= viaje_admin_form
    success_url= "viajes_admin"

    
@premiso_admin
def listar_viajes_admin(request):
    
    numero_viaje_exacto = request.GET.get("dato")
    qs = ViajeAdministrativo.objects.all()

    if numero_viaje_exacto != "" and numero_viaje_exacto is not None:
        qs = qs.filter(numero_viaje__icontains=numero_viaje_exacto)

    context = {
        "object_list": qs,
    }

    return render(request, "viajes_templates/listar_viajes_administrativos.html", context)