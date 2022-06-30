from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from apps.home.decoradores_viaje import permiso_transportista, permiso_staff
from django.views.generic import CreateView, View, ListView, TemplateView
from django.core import serializers
from django.contrib import messages
import pytz
from apps.account.models import Account as User
from apps.viajes.models import Viaje

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from datetime import datetime
import random
from django.db.models import Count, Sum, Avg, Q


# Create your views here.


fecha_tiempo = datetime.now()
hora_actual = str(fecha_tiempo.hour)
minuto_actual = str(fecha_tiempo.minute)
dia = str(fecha_tiempo.day)
mes = str(fecha_tiempo.month)
ano = str(fecha_tiempo.year)
 
# esta funcion devuelve los viajes que se han realizado por el transportista logueado al momento
@permiso_staff
@login_required
def reporte_tr(request):
 
    numero_exacto_viaje = request.GET.get("dato")
    #en Sqlite la funcion Distict no funciona con parametros, solo funciona en postgress, por tal razon si se usa Sqlite esta funcion dara error, pero para traer valores distintos en base al numero de viaje se debe hacer asi
    #en Sqlite distinct solo funciona con un solo value y sin parametros en la funcion distinct, se emplo asi aqui porque en heroku usamos postgresSQL
    #la linea comentada, funcionara con Sqlite
    
    #qs = Viaje.objects.values('id_viaje','numero_viaje','transportista','fecha_viaje','hora_viaje', 'tipo_viaje').distinct('numero_viaje').order_by('-numero_viaje')
    qs = Viaje.objects.values('id_viaje','numero_viaje','transportista','fecha_viaje','hora_viaje', 'tipo_viaje').distinct('numero_viaje').order_by('-numero_viaje')

    if numero_exacto_viaje != "" and numero_exacto_viaje is not None:
        qs = qs.filter(
            
            numero_viaje__icontains=numero_exacto_viaje,
        ).order_by('-fecha_viaje')

    context = {
        "object_list": qs,
    }

    return render(request, "reportes_templates/reporte_viajes_t.html", context)


# esta funcion devuelve los viajes con sus precios, tomando como filtro el numero de viaje que que esta en el momento

@permiso_staff
def detalle_viaje(request, numero_viaje):
    
    try:
        qs = Viaje.objects.filter(numero_viaje=numero_viaje).order_by('-fecha_viaje')
        # contar el numero de pasajeros en este queryset
        cantidad_pasajeros = qs.count()

        # definir el precio del viaje
        # viajes de 1 a 3 personas paga 12 dolares
        # viajes de 4 en adelante personas paga 26 dolares
        
        if cantidad_pasajeros >= 1 and cantidad_pasajeros < 4:
            precio_viaje = 12

        if cantidad_pasajeros >= 4 :
            precio_viaje = 26

        if cantidad_pasajeros < 1:
            precio_viaje = 0

        # calcular el precio por cada pasajero
        if cantidad_pasajeros > 0:
            precio_por_pasajero = precio_viaje / cantidad_pasajeros
        else:
            precio_por_pasajero = 0

        context = {
            "object_list": qs,
            "cantidad_pasajeros": cantidad_pasajeros,
            "precio_viaje": precio_viaje,
            "precio_por_pasajero": precio_por_pasajero,
        }

        return render(request, "reportes_templates/detalle_viaje.html", context)
    except Exception as e:
        messages.success(request, str(e) +' '+ "A ocurrido un Error al cargar los detalles del viaje")
        return redirect('home')



@method_decorator(permiso_staff,name='dispatch')
@method_decorator(login_required, name='dispatch')
class reporte_viaje_excel(TemplateView):

    # Usamos el método get para generar el archivo excel
      # Obtenemos todas los viajes  de nuestra base de datos o lo los filtramos
    def get(self, request, *args, **kwargs):
        numero_exacto_viaje = request.GET.get("dato")
        viajes = Viaje.objects.all().order_by('-fecha_viaje')
        if numero_exacto_viaje != "" and numero_exacto_viaje is not None:
            viajes = viajes.filter(
                numero_viaje__icontains=numero_exacto_viaje,
            ).order_by('-fecha_viaje')

       
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        # Aplicamos estilo a la columna B1
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B2"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["B2"].fill = PatternFill(
            start_color="66FFCC", end_color="66FFCC", fill_type="solid"
        )
        ws["B2"].font = Font(name="Calibi", size=16, bold=True)
        ws["B2"] = "REPORTE DE VIAJES"
        # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells("B2:P2")
        ws.row_dimensions[2].height = 25

        # definimos el tamaño de cada colunma algunas las dejamos en automatico
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["d"].auto_size = True
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].auto_size = True
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].auto_size = True
        ws.column_dimensions["J"].auto_size = True
        ws.column_dimensions["K"].auto_size = True
        ws.column_dimensions["L"].auto_size = True
        ws.column_dimensions["M"].width = 15
        ws.column_dimensions["N"].width = 30
        ws.column_dimensions["O"].width = 30
        ws.column_dimensions["P"].auto_size = True

        # Creamos los encabezados desde la celda B3 hasta la E3
        ws["B5"] = "Id"
        ws["C5"] = "Numero de viaje"
        ws["D5"] = "Transportista"
        ws["E5"] = "Fecha"
        ws["F5"] = "Hora"
        ws["G5"] = "Tipo"
        ws["H5"] = "Id"
        ws["I5"] = "Nombre"
        ws["J5"] = "Apellido"
        ws["K5"] = "Campaña"
        ws["L5"] = "Site Pasajero"
        ws["M5"] = "Ruta"
        ws["N5"] = "Direccion"
        ws["O5"] = "Excepcion"
        ws["P5"] = "Site Ruta"
        # ws['P3']='Hora de entrada del pasajero'
        # ws['P3']='Hora de salida del pasajero'

        # damos Formato a cada una de las celdas utilizadas
        ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["B5"].font = Font(name="Arial", size=10, bold=True)
        ws["B5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["C5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["C5"].font = Font(name="Arial", size=10, bold=True)
        ws["C5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["D5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["D5"].font = Font(name="Arial", size=10, bold=True)
        ws["D5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["E5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["E5"].font = Font(name="Arial", size=10, bold=True)
        ws["E5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["F5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["F5"].font = Font(name="Arial", size=10, bold=True)
        ws["F5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["G5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["G5"].font = Font(name="Arial", size=10, bold=True)
        ws["G5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["H5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["H5"].font = Font(name="Arial", size=10, bold=True)
        ws["H5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["I5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["I5"].font = Font(name="Arial", size=10, bold=True)
        ws["I5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["J5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["J5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["J5"].font = Font(name="Arial", size=10, bold=True)
        ws["J5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["K5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["K5"].font = Font(name="Arial", size=10, bold=True)
        ws["K5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["L5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["L5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["L5"].font = Font(name="Arial", size=10, bold=True)
        ws["L5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["M5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["M5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["M5"].font = Font(name="Arial", size=10, bold=True)
        ws["M5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["N5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["N5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["N5"].font = Font(name="Arial", size=10, bold=True)
        ws["N5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["O5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["O5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["O5"].font = Font(name="Arial", size=10, bold=True)
        ws["O5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        ws["P5"].alignment = Alignment(horizontal="center", vertical="center")
        ws["P5"].border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin"),
        )
        ws["P5"].font = Font(name="Arial", size=10, bold=True)
        ws["P5"].fill = PatternFill(
            start_color="66CFCC", end_color="66CFCC", fill_type="solid"
        )

        

        cont = 6
        # Recorremos el conjunto de viajes y vamos escribiendo cada uno de los datos en las celdas y tambien se le asignan su respectivos estilos
        for viaje in viajes:
            ws.cell(row=cont, column=2).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=2).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=2).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=2).value = str(viaje.id_viaje)

            ws.cell(row=cont, column=3).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=3).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=3).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=3).value = str(viaje.numero_viaje)

            ws.cell(row=cont, column=4).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=4).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=4).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=4).value = str(viaje.transportista)

            ws.cell(row=cont, column=5).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=5).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=5).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=5).value = str(viaje.fecha_viaje)

            ws.cell(row=cont, column=6).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=6).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=6).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=6).value = str(viaje.hora_viaje)

            ws.cell(row=cont, column=7).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=7).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=7).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=7).value = str(viaje.tipo_viaje)

            ws.cell(row=cont, column=8).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=8).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=8).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=8).value = str(viaje.id_pasajero)

            ws.cell(row=cont, column=9).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=9).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=9).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=9).value = str(viaje.nombre_pasajero)

            ws.cell(row=cont, column=10).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=10).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=10).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=10).value = str(viaje.apellido_pasajero)

            ws.cell(row=cont, column=11).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=11).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=11).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=11).value = str(viaje.campaña_pasajero)

            ws.cell(row=cont, column=12).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=12).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=12).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=12).value = str(viaje.site_pasajero)

            ws.cell(row=cont, column=13).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=13).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=13).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=13).value = str(viaje.ruta_pasajero)

            ws.cell(row=cont, column=14).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=14).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=14).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=14).value = str(viaje.direccion_pasajero)

            ws.cell(row=cont, column=15).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=15).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=15).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=15).value = str(viaje.razon_excepcion)

            ws.cell(row=cont, column=16).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=cont, column=16).border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )
            ws.cell(row=cont, column=16).font = Font(name="Arial", size=8, bold=True)
            ws.cell(row=cont, column=16).value = str(viaje.site_destino_origen)

            # ws.cell(row = cont, column =17).value = viaje.hora_entrada
            # ws.cell(row = cont, column =18).value = viaje.hora_salida
            cont = cont + 1
        # Establecemos el nombre del archivo este estara compuesto por la palabra TPGP-REPORT- +dia + mes+ ano +horaactual+minutoactual
        nombre_archivo = (
            "TPGO-REPORT-" + dia + mes + ano + hora_actual + minuto_actual + ".xlsx"
        )
        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

